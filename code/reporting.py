"""Utilities for exporting formatted Excel reports."""
from __future__ import annotations

import re
from collections import Counter
from pathlib import Path
from typing import Dict, List, Mapping, Optional, Tuple

import pandas as pd
from openpyxl.styles import Alignment, Border, PatternFill, Side, Font
from openpyxl.utils import get_column_letter

from config import Config
from response_parser import ResponseParser

# Shared column identifiers used across batch_analyzer/reporting.
AI_SUCCESS_COUNT_COLUMN = "AI Successful Runs"
AI_TOTAL_COUNT_COLUMN = "AI Total Runs"
AI_SUCCESS_RATE_COLUMN = "AI Success Rate"
Q15_VOTE_COUNTS_COLUMN = "Vote Counts"
HUMAN_Q15_TYPE_COLUMN = "Human Type"
AI_Q15_TYPE_COLUMN = "AI Majority Type"
MISMATCH_PAIR_COLUMN = "Mismatch Pair (Human→AI)"
DETAIL_NOTE_COLUMN = "Detail Note"

# Pre-compute question column naming helpers.
QUESTION_COLUMN_PATTERN = re.compile(r"\[Q(\d+)\]")


STATUS_COLORS = {
    "Pass_Strong": PatternFill("solid", fgColor="93C47D"),  # deep green
    "Pass_Weak": PatternFill("solid", fgColor="C6EFCE"),  # light green
    "Ambiguous_Tie": PatternFill("solid", fgColor="F9CB9C"),  # orange
    "Ambiguous_PoorCoverage": PatternFill("solid", fgColor="FCE5CD"),  # light orange
    "Contradiction": PatternFill("solid", fgColor="F4CCCC"),  # red
    "Technical_Failure": PatternFill("solid", fgColor="D9D9D9"),  # gray
}

STATUS_DESCRIPTIONS = {
    "Pass_Strong": "Human label matches AI majority; AI runs unanimous or strong majority.",
    "Pass_Weak": "Human matches AI majority, but consensus strength is lower (simple majority).",
    "Contradiction": "Human label conflicts with AI majority decision.",
    "Ambiguous_Tie": "AI runs tie or split; majority outcome indeterminate.",
    "Ambiguous_PoorCoverage": "Insufficient consensus (missing data, plurality, single run, or no majority).",
    "Technical_Failure": "No successful AI runs (PDF/read/analysis failures).",
}


def export_excel(
    df: pd.DataFrame,
    output_path: Path,
    *,
    article_id_column: str = "#",
    source_column: str = "source",
    title_column: str = "Title of the Paper",
    article_status_column: str = "Article_Status",
    ai_agreement_column: str = "AI run agreement (Q15)",
    human_vs_ai_column: str = "Human vs AI (Q15)",
    human_vs_consensus_column: str = "Human vs AI (consensus)",
    type_summary_column: str = "Type summary (Q15, Decision Tree, Consensus)",
) -> pd.DataFrame:
    """Export analysis results to a richly formatted Excel workbook."""
    if df.empty:
        output_path.parent.mkdir(parents=True, exist_ok=True)
        with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
            df.to_excel(writer, sheet_name="All_Results", index=False)
        return pd.DataFrame(columns=[article_id_column, article_status_column])

    df_all = df.copy()
    summary_df, status_map, run_stats = _build_article_summary(
        df_all,
        article_id_column,
        source_column,
        title_column,
        article_status_column,
        ai_agreement_column,
        human_vs_ai_column,
        human_vs_consensus_column,
        type_summary_column,
    )

    status_series = df_all[article_id_column].apply(lambda value: status_map.get(str(value), ""))
    if article_status_column in df_all.columns:
        df_all[article_status_column] = status_series
    else:
        insert_at = 2 if "Analysis_Status" in df_all.columns else 1
        df_all.insert(insert_at, article_status_column, status_series)

    detail_frames, detail_template = _build_status_detail_frames(
        df_all,
        summary_df,
        article_id_column,
        source_column,
        article_status_column,
        ai_agreement_column,
        human_vs_ai_column,
        type_summary_column,
    )

    summary_tables = _build_summary_tables(
        summary_df,
        article_status_column,
        ai_agreement_column,
        human_vs_ai_column,
    )

    analytics_tables = _build_analytics_tables(
        df_all,
        summary_df,
        article_status_column,
        ai_agreement_column,
        human_vs_ai_column,
    )

    output_path = Path(output_path)
    if output_path.suffix.lower() != ".xlsx":
        output_path = output_path.with_suffix(".xlsx")
    output_path.parent.mkdir(parents=True, exist_ok=True)

    status_order = [
        "Pass_Strong",
        "Pass_Weak",
        "Contradiction",
        "Ambiguous_Tie",
        "Ambiguous_PoorCoverage",
        "Technical_Failure",
    ]
    observed_statuses = summary_df[article_status_column].dropna().unique().tolist()
    remaining_statuses = [status for status in observed_statuses if status not in status_order]
    ordered_statuses = status_order + sorted(remaining_statuses)
    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        df_all.to_excel(writer, sheet_name="All_Results", index=False)
        summary_df.to_excel(writer, sheet_name="Article_Summary", index=False)

        for status in ordered_statuses:
            detail_df = detail_frames.get(status)
            if detail_df is None or detail_df.empty:
                detail_df = detail_template.copy()
            detail_df.to_excel(writer, sheet_name=status, index=False)

        sheets = writer.sheets

        _format_all_results_sheet(
            sheets.get("All_Results"),
            article_id_column,
            article_status_column,
        )
        _format_summary_sheet(
            sheets.get("Article_Summary"),
            article_status_column,
        )
        for status in ordered_statuses:
            _format_all_results_sheet(
                sheets.get(status),
                article_id_column,
                article_status_column,
            )

        workbook = writer.book
        _populate_summary_sheet(
            workbook,
            summary_tables,
        )
        for sheet_name, tables in analytics_tables.items():
            _populate_custom_sheet(
                workbook,
                sheet_name,
                tables,
            )

    _print_status_report(status_map, run_stats)
    return summary_df


def _build_article_summary(
    df: pd.DataFrame,
    article_id_column: str,
    source_column: str,
    title_column: str,
    article_status_column: str,
    ai_agreement_column: str,
    human_vs_ai_column: str,
    human_vs_consensus_column: str,
    type_summary_column: str,
) -> Tuple[pd.DataFrame, Dict[str, str], Dict[str, Counter]]:
    summary_rows: List[Dict[str, object]] = []
    status_map: Dict[str, str] = {}
    run_stats: Dict[str, Counter] = {}
    question_map = _build_question_column_map(df.columns)
    type_ids = sorted(Config.TYPE_QUESTION_GROUPS.keys())

    for article_id, group in df.groupby(article_id_column, sort=True):
        human_rows = group[group[source_column] == "human"]
        human_row = human_rows.iloc[0] if not human_rows.empty else group.iloc[0]

        majority_rows = group[group[source_column].astype(str).str.contains("majority-vote", na=False)]
        majority_row = majority_rows.iloc[0] if not majority_rows.empty else None

        ai_rows = group[group[source_column] != "human"].copy()
        ai_base_rows = ai_rows[~ai_rows[source_column].astype(str).str.contains("majority-vote", na=False)]
        success_count = ai_base_rows["Analysis_Status"].astype(str).str.startswith("SUCCESS_").sum()
        total_ai_runs = len(ai_base_rows)
        has_success = success_count > 0
        has_majority = ai_rows[source_column].astype(str).str.contains("majority-vote", na=False).any()

        human_vs_ai = str(human_row.get(human_vs_ai_column, "") or "")
        human_vs_consensus = str(human_row.get(human_vs_consensus_column, "") or "")
        ai_agreement_label = str(human_row.get(ai_agreement_column, "") or "")

        # Determine columns for Human (Q15) and AI (Q16) classification
        q_human_col = question_map.get(15)
        q_ai_col = question_map.get(Config.Q_ID_CLASSIFICATION)

        human_q15 = str(human_row.get(q_human_col, "") or "").strip() if q_human_col else ""
        ai_majority_q15 = ""
        if majority_row is not None and q_ai_col:
            ai_majority_q15 = str(majority_row.get(q_ai_col, "") or "").strip()
        if not ai_majority_q15:
            agreement_lower = ai_agreement_label.lower()
            human_vs_ai_lower = human_vs_ai.lower()
            if "tie" in human_vs_ai_lower or "split consensus" in agreement_lower:
                ai_majority_q15 = "Tie (no majority)"

        vote_counts_text = ""
        if majority_row is not None:
            vote_counts_text = str(majority_row.get(Q15_VOTE_COUNTS_COLUMN, "") or "")
        if not vote_counts_text:
            vote_counts_counter = _derive_q15_vote_counter(ai_base_rows, q_ai_col)
            vote_counts_text = _format_vote_counter(vote_counts_counter)
        if not ai_majority_q15 and vote_counts_text.lower().startswith("tie"):
            ai_majority_q15 = "Tie (no majority)"

        success_rate = None
        success_rate_raw = human_row.get(AI_SUCCESS_RATE_COLUMN)
        if success_rate_raw not in ("", None):
            try:
                success_rate = float(success_rate_raw)
            except (TypeError, ValueError):
                success_rate = None
        if success_rate is None and total_ai_runs:
            success_rate = round(success_count / total_ai_runs, 3)

        extent_avgs, confidence_avgs = _collect_type_metrics(ai_base_rows, question_map)

        status = _determine_article_status(
            human_vs_ai,
            human_vs_consensus,
            has_success,
            ai_agreement_label=ai_agreement_label,
            majority_available=has_majority,
            vote_counts_text=vote_counts_text,
            success_count=success_count,
            total_runs=total_ai_runs,
            ai_majority_value=ai_majority_q15,
        )
        status_map[str(article_id)] = status

        summary_row = {
            article_id_column: article_id,
            article_status_column: status,
            title_column: human_row.get(title_column, human_row.get("Title", "")),
            HUMAN_Q15_TYPE_COLUMN: human_q15,
            AI_Q15_TYPE_COLUMN: ai_majority_q15,
            ai_agreement_column: ai_agreement_label,
            human_vs_ai_column: human_vs_ai,
            human_vs_consensus_column: human_vs_consensus,
            type_summary_column: human_row.get(type_summary_column, ""),
            AI_SUCCESS_COUNT_COLUMN: int(success_count),
            AI_TOTAL_COUNT_COLUMN: int(total_ai_runs),
            AI_SUCCESS_RATE_COLUMN: success_rate,
            "Majority Vote Available": "Yes" if has_majority else "No",
            Q15_VOTE_COUNTS_COLUMN: vote_counts_text,
        }

        for type_id in type_ids:
            confidence_col_name = f"Type {type_id} confidence avg"
            confidence_value = confidence_avgs.get(type_id)
            summary_row[confidence_col_name] = round(confidence_value, 3) if confidence_value is not None else None

        summary_rows.append(summary_row)

        outcome_counter = Counter()
        for status_value in ai_base_rows["Analysis_Status"].astype(str):
            key = status_value.split("_")[0] if status_value else "UNKNOWN"
            outcome_counter[key] += 1
        run_stats[str(article_id)] = outcome_counter

    summary_df = pd.DataFrame(summary_rows)
    if not summary_df.empty:
        summary_df = summary_df.sort_values(article_id_column).reset_index(drop=True)

    desired_order = [
        article_id_column,
        title_column,
        article_status_column,
        HUMAN_Q15_TYPE_COLUMN,
        AI_Q15_TYPE_COLUMN,
        ai_agreement_column,
        Q15_VOTE_COUNTS_COLUMN,
        human_vs_ai_column,
        human_vs_consensus_column,
        type_summary_column,
        AI_SUCCESS_COUNT_COLUMN,
        AI_TOTAL_COUNT_COLUMN,
        AI_SUCCESS_RATE_COLUMN,
        "Majority Vote Available",
    ]
    for type_id in type_ids:
        desired_order.append(f"Type {type_id} confidence avg")
    existing_columns = [col for col in desired_order if col in summary_df.columns]
    remaining_columns = [col for col in summary_df.columns if col not in existing_columns]
    summary_df = summary_df[existing_columns + remaining_columns]
    return summary_df, status_map, run_stats


def _build_status_detail_frames(
    df_all: pd.DataFrame,
    summary_df: pd.DataFrame,
    article_id_column: str,
    source_column: str,
    article_status_column: str,
    ai_agreement_column: str,
    human_vs_ai_column: str,
    type_summary_column: str,
) -> Tuple[Dict[str, pd.DataFrame], pd.DataFrame]:
    """Create per-status detail tables containing human, majority, and run rows."""
    detail_frames: Dict[str, pd.DataFrame] = {}
    summary_lookup = summary_df.set_index(article_id_column, drop=False)
    base_columns = list(df_all.columns)
    for extra_col in [MISMATCH_PAIR_COLUMN, DETAIL_NOTE_COLUMN]:
        if extra_col not in base_columns:
            base_columns.append(extra_col)

    for status_value, status_rows in summary_df.groupby(article_status_column):
        article_ids = status_rows[article_id_column].tolist()
        status_details: List[pd.DataFrame] = []
        for article_id in article_ids:
            article_rows = df_all[df_all[article_id_column] == article_id].copy()
            if article_rows.empty:
                continue

            summary_row = summary_lookup.loc[article_id]
            if isinstance(summary_row, pd.DataFrame):
                summary_row = summary_row.iloc[0]

            prepared = _prepare_article_detail_rows(
                article_rows,
                summary_row,
                source_column,
                ai_agreement_column,
                human_vs_ai_column,
                type_summary_column,
                base_columns,
            )
            status_details.append(prepared)

        if status_details:
            detail_frames[status_value] = pd.concat(status_details, ignore_index=True)
        else:
            detail_frames[status_value] = pd.DataFrame(columns=base_columns)

    template_df = pd.DataFrame(columns=base_columns)
    return detail_frames, template_df


def _build_summary_tables(
    summary_df: pd.DataFrame,
    article_status_column: str,
    ai_agreement_column: str,
    human_vs_ai_column: str,
) -> List[Tuple[str, pd.DataFrame]]:
    total_articles = len(summary_df)
    summary_copy = summary_df.copy()

    # Numeric conversions
    success_rates = pd.to_numeric(summary_copy.get(AI_SUCCESS_RATE_COLUMN), errors="coerce")
    total_runs_series = pd.to_numeric(summary_copy.get(AI_TOTAL_COUNT_COLUMN), errors="coerce")
    success_count_series = pd.to_numeric(summary_copy.get(AI_SUCCESS_COUNT_COLUMN), errors="coerce")

    majority_available = summary_copy.get("Majority Vote Available", pd.Series([], dtype=str))
    majority_yes = int((majority_available.astype(str).str.upper() == "YES").sum())

    successful_articles = int((success_count_series.fillna(0) > 0).sum()) if not success_count_series.empty else 0

    match_mask = summary_copy.get(human_vs_ai_column, pd.Series([], dtype=str)).fillna("").str.startswith("Match", na=False)
    comparison_mask = summary_copy.get(human_vs_ai_column, pd.Series([], dtype=str)).fillna("").str.contains(
        "match", case=False, na=False
    )
    accuracy_count = int(match_mask.sum())
    accuracy_denominator = int(comparison_mask.sum())

    status_counts = summary_copy.get(article_status_column, pd.Series([], dtype=str)).fillna("Unknown").value_counts()
    status_counts = status_counts.reindex(
        ["Pass_Strong", "Pass_Weak", "Contradiction", "Ambiguous_Tie", "Ambiguous_PoorCoverage", "Technical_Failure"],
        fill_value=0,
    )
    ambiguous_count = int(
        status_counts.get("Ambiguous_Tie", 0) + status_counts.get("Ambiguous_PoorCoverage", 0)
    )
    technical_count = int(status_counts.get("Technical_Failure", 0))
    contradiction_count = int(status_counts.get("Contradiction", 0))

    average_success_rate = float(success_rates.mean()) if not success_rates.dropna().empty else None
    average_runs = float(total_runs_series.mean()) if not total_runs_series.dropna().empty else None

    metrics_rows = [
        {"Metric": "Total articles", "Value": str(total_articles)},
        {
            "Metric": "AI majority coverage",
            "Value": _format_ratio(majority_yes, total_articles),
        },
        {
            "Metric": "Articles with ≥1 successful AI run",
            "Value": _format_ratio(successful_articles, total_articles),
        },
        {
            "Metric": "Accuracy vs human (Q15)",
            "Value": _format_ratio(accuracy_count, accuracy_denominator),
        },
        {
            "Metric": "Ambiguous outcomes",
            "Value": _format_ratio(ambiguous_count, total_articles),
        },
        {
            "Metric": "Contradictions",
            "Value": _format_ratio(contradiction_count, total_articles),
        },
        {
            "Metric": "Technical failures",
            "Value": _format_ratio(technical_count, total_articles),
        },
        {
            "Metric": "Average AI success rate",
            "Value": _format_percentage(average_success_rate),
        },
        {
            "Metric": "Average AI runs per article",
            "Value": f"{average_runs:.1f}" if average_runs is not None else "N/A",
        },
    ]
    metrics_df = pd.DataFrame(metrics_rows)

    status_breakdown_rows = []
    for status, count in status_counts.items():
        status_breakdown_rows.append(
            {
                "Status": status,
                "Meaning": STATUS_DESCRIPTIONS.get(status, ""),
                "Articles": count,
                "Share": _format_ratio(count, total_articles),
            }
        )
    status_breakdown_df = pd.DataFrame(status_breakdown_rows)

    agreement_series = summary_copy.get(ai_agreement_column, pd.Series([], dtype=str)).fillna("Missing")
    agreement_counts = agreement_series.value_counts()
    total_agreement = int(agreement_counts.sum())
    agreement_rows = []
    for label, count in agreement_counts.items():
        agreement_rows.append(
            {
                "Agreement strength": label,
                "Articles": int(count),
                "Share": _format_ratio(int(count), total_agreement),
            }
        )
    agreement_df = pd.DataFrame(agreement_rows)

    glossary_rows = [
        {
            "Term": "AI majority coverage",
            "Description": "Articles where a majority vote was available (at least two successful runs).",
        },
        {
            "Term": "Accuracy vs human (Q15)",
            "Description": "Share of articles where the AI majority Q15 matches the human label among cases with match/mismatch labels.",
        },
        {
            "Term": "AI run agreement (Q15)",
            "Description": "Consensus strength reported on majority rows (Unanimous, Strong majority, etc.).",
        },
        {
            "Term": "Detail Note",
            "Description": "In status sheets, captures context for each row (human label, consensus outcome, or run status).",
        },
        {
            "Term": "Mismatch Pair (Human→AI)",
            "Description": "For contradictions, records the human 4PT type and the AI majority type.",
        },
    ]
    glossary_df = pd.DataFrame(glossary_rows)

    legend_rows = []
    for status, description in STATUS_DESCRIPTIONS.items():
        legend_rows.append(
            {
                "Status": status,
                "Meaning": description,
            }
        )
    legend_df = pd.DataFrame(legend_rows)

    tables = [
        ("Summary Metrics", metrics_df),
        ("Status Breakdown", status_breakdown_df),
        ("Agreement Breakdown", agreement_df),
        ("Terminology", glossary_df),
        ("Status Colour Legend", legend_df),
    ]
    return tables


def _build_analytics_tables(
    df_all: pd.DataFrame,
    summary_df: pd.DataFrame,
    article_status_column: str,
    ai_agreement_column: str,
    human_vs_ai_column: str,
) -> Dict[str, List[Tuple[str, pd.DataFrame]]]:
    tables: Dict[str, List[Tuple[str, pd.DataFrame]]] = {}

    confusion_tables = _build_confusion_tables(summary_df)
    if confusion_tables:
        tables["Confusion_Matrix"] = confusion_tables

    agreement_tables = _build_agreement_tables(summary_df, ai_agreement_column)
    if agreement_tables:
        tables["Agreement_Distribution"] = agreement_tables

    ambiguity_tables = _build_ambiguity_tables(summary_df, article_status_column, ai_agreement_column)
    if ambiguity_tables:
        tables["Ambiguity_Patterns"] = ambiguity_tables

    margin_tables = _build_margin_tables(summary_df, human_vs_ai_column)
    if margin_tables:
        tables["Majority_Margin"] = margin_tables

    return tables


def _build_confusion_tables(summary_df: pd.DataFrame) -> List[Tuple[str, pd.DataFrame]]:
    if summary_df.empty:
        return []

    human_series = summary_df.get(HUMAN_Q15_TYPE_COLUMN)
    ai_series = summary_df.get(AI_Q15_TYPE_COLUMN)
    if human_series is None or ai_series is None:
        return []

    human_series = human_series.fillna("Unknown")
    ai_series = ai_series.fillna("No majority")

    if human_series.empty:
        return []

    tables: List[Tuple[str, pd.DataFrame]] = []

    # 多分类混淆矩阵
    tables.extend(_build_confusion_matrix_tables(human_series, ai_series, title_prefix=""))

    # 二分类混淆矩阵：Type1/2 vs Type3/4（其余归为 Other/Unknown）
    tables.extend(_build_binary_confusion_tables(human_series, ai_series))

    return tables


def _build_binary_confusion_tables(
    human_series: pd.Series,
    ai_series: pd.Series,
) -> List[Tuple[str, pd.DataFrame]]:
    def _bin_label(label: str) -> str:
        text = str(label or "").strip().lower()
        if text.startswith("type 1") or text == "1":
            return "Type1/2"
        if text.startswith("type 2") or text == "2":
            return "Type1/2"
        if text.startswith("type 3") or text == "3":
            return "Type3/4"
        if text.startswith("type 4") or text == "4":
            return "Type3/4"
        return "Other/Unknown"

    human_bin = human_series.apply(_bin_label)
    ai_bin = ai_series.apply(_bin_label)

    labels = sorted(set(human_bin.unique()).union(set(ai_bin.unique())))
    if len(labels) <= 1:
        return []

    return _build_confusion_matrix_tables(human_bin, ai_bin, title_prefix="Binary (Type1+2 vs Type3+4) - ")


def _build_confusion_matrix_tables(
    human_series: pd.Series,
    ai_series: pd.Series,
    title_prefix: str = "",
) -> List[Tuple[str, pd.DataFrame]]:
    labels = sorted(set(human_series.unique()).union(set(ai_series.unique())))

    counts = pd.crosstab(human_series, ai_series, dropna=False).reindex(index=labels, columns=labels, fill_value=0)

    def _format_percent_df(df: pd.DataFrame) -> pd.DataFrame:
        formatted = df.copy()
        for col in formatted.columns:
            formatted[col] = formatted[col].apply(lambda x: f"{x * 100:.1f}%")
        formatted.index.name = df.index.name
        formatted.columns.name = df.columns.name
        return formatted

    row_sum = counts.sum(axis=1).replace(0, pd.NA)
    row_norm = counts.div(row_sum, axis=0).fillna(0.0)
    col_sum = counts.sum(axis=0).replace(0, pd.NA)
    col_norm = counts.div(col_sum, axis=1).fillna(0.0)

    metrics_rows = []
    recalls = []
    f1_scores = []
    for label in labels:
        tp = counts.at[label, label] if label in counts.columns else 0
        actual = counts.loc[label].sum()
        predicted = counts[label].sum() if label in counts.columns else 0
        recall = tp / actual if actual else None
        precision = tp / predicted if predicted else None
        if precision is not None and recall is not None and (precision + recall) > 0:
            f1 = 2 * precision * recall / (precision + recall)
        else:
            f1 = None
        metrics_rows.append(
            {
                "Type": label,
                "Support (human)": actual,
                "Predicted (AI)": predicted,
                "Precision": _format_percentage(precision) if precision is not None else "N/A",
                "Recall": _format_percentage(recall) if recall is not None else "N/A",
                "F1": _format_percentage(f1) if f1 is not None else "N/A",
            }
        )
        if recall is not None:
            recalls.append(recall)
        if f1 is not None:
            f1_scores.append(f1)

    macro_recall = sum(recalls) / len(recalls) if recalls else None
    macro_f1 = sum(f1_scores) / len(f1_scores) if f1_scores else None
    metrics_rows.append(
        {
            "Type": "Macro Average",
            "Support (human)": counts.values.sum(),
            "Predicted (AI)": counts.values.sum(),
            "Precision": "—",
            "Recall": _format_percentage(macro_recall) if macro_recall is not None else "N/A",
            "F1": _format_percentage(macro_f1) if macro_f1 is not None else "N/A",
        }
    )

    top_errors_rows = []
    for human_label in labels:
        for ai_label in labels:
            if human_label == ai_label:
                continue
            count = counts.at[human_label, ai_label]
            if count > 0:
                top_errors_rows.append({"Human Type": human_label, "AI Type": ai_label, "Count": count})
    top_errors_df = pd.DataFrame(top_errors_rows)
    if not top_errors_df.empty:
        top_errors_df = top_errors_df.sort_values("Count", ascending=False).reset_index(drop=True)

    tables = [
        (
            f"{title_prefix}Confusion matrix (counts)",
            counts.reset_index().rename(columns={counts.index.name or "row_0": "Human Type"}),
        ),
        (
            f"{title_prefix}Row-normalised (%)",
            _format_percent_df(row_norm).reset_index().rename(columns={row_norm.index.name or "index": "Human Type"}),
        ),
        (
            f"{title_prefix}Column-normalised (%)",
            _format_percent_df(col_norm).reset_index().rename(columns={col_norm.index.name or "index": "Human Type"}),
        ),
        (f"{title_prefix}Per-class metrics", pd.DataFrame(metrics_rows)),
    ]
    if not top_errors_df.empty:
        tables.append((f"{title_prefix}Top error pairs", top_errors_df))
    return tables


def _build_agreement_tables(summary_df: pd.DataFrame, ai_agreement_column: str) -> List[Tuple[str, pd.DataFrame]]:
    agreement_series = summary_df.get(ai_agreement_column)
    if agreement_series is None or agreement_series.empty:
        return []
    counts = agreement_series.fillna("Missing").value_counts().reset_index()
    counts.columns = ["Agreement strength", "Articles"]
    total = counts["Articles"].sum()
    counts["Share"] = counts["Articles"].apply(lambda x: _format_ratio(int(x), int(total)))
    return [("AI run agreement distribution", counts)]


def _build_ambiguity_tables(
    summary_df: pd.DataFrame,
    article_status_column: str,
    ai_agreement_column: str,
) -> List[Tuple[str, pd.DataFrame]]:
    if article_status_column not in summary_df.columns:
        return []
    ambiguous_mask = summary_df[article_status_column].isin(["Ambiguous_Tie", "Ambiguous_PoorCoverage"])
    ambiguous_df = summary_df[ambiguous_mask].copy()
    if ambiguous_df.empty:
        return []

    patterns = []
    for _, row in ambiguous_df.iterrows():
        vote_text = str(row.get(Q15_VOTE_COUNTS_COLUMN, "") or "")
        pattern_key, top_votes, second_votes = _summarise_vote_pattern(vote_text)
        patterns.append(
            {
                "Article #": row.get("#"),
                "Article_Status": row.get(article_status_column),
                "AI run agreement (Q15)": row.get(ai_agreement_column),
                "Vote counts": vote_text,
                "Pattern": pattern_key,
                "Margin": top_votes - second_votes,
            }
        )
    pattern_df = pd.DataFrame(patterns)

    pattern_summary = (
        pattern_df.groupby("Pattern")
        .size()
        .reset_index(name="Articles")
        .sort_values("Articles", ascending=False)
    )
    total = pattern_summary["Articles"].sum()
    pattern_summary["Share"] = pattern_summary["Articles"].apply(lambda x: _format_ratio(int(x), int(total)))

    return [
        ("Ambiguous vote patterns", pattern_summary),
        ("Ambiguous article details", pattern_df),
    ]


def _build_margin_tables(summary_df: pd.DataFrame, human_vs_ai_column: str) -> List[Tuple[str, pd.DataFrame]]:
    if summary_df.empty:
        return []
    margin_rows = []
    for _, row in summary_df.iterrows():
        vote_text = str(row.get(Q15_VOTE_COUNTS_COLUMN, "") or "")
        pattern_key, top_votes, second_votes = _summarise_vote_pattern(vote_text)
        total_votes = top_votes + second_votes
        margin = top_votes - second_votes
        normalized_margin = (margin / total_votes) if total_votes else None
        match_label = str(row.get(human_vs_ai_column, "") or "")
        is_match = match_label.lower().startswith("match")
        margin_rows.append(
            {
                "Article #": row.get("#"),
                "Human vs AI (Q15)": match_label,
                "Vote counts": vote_text,
                "Margin": margin,
                "Normalised margin": normalized_margin,
                "Pattern": pattern_key,
                "Match": "Yes" if is_match else "No",
            }
        )

    margin_df = pd.DataFrame(margin_rows)
    if margin_df.empty:
        return []

    def _bucket_margin(value: int) -> str:
        if value <= 0:
            return "0 (tie)"
        if value == 1:
            return "1"
        return "≥2"

    margin_df["Margin bucket"] = margin_df["Margin"].apply(lambda x: _bucket_margin(int(x) if pd.notna(x) else 0))

    bucket_summary = (
        margin_df.groupby("Margin bucket")
        .agg(Articles=("Article #", "count"), Matches=("Match", lambda s: (s == "Yes").sum()))
        .reset_index()
    )
    bucket_summary["Accuracy"] = bucket_summary.apply(
        lambda row: _format_ratio(int(row["Matches"]), int(row["Articles"])), axis=1
    )

    return [
        ("Margin buckets", bucket_summary[["Margin bucket", "Articles", "Accuracy"]]),
        ("Article-level margins", margin_df),
    ]


def _populate_summary_sheet(workbook, tables: List[Tuple[str, pd.DataFrame]]) -> None:
    if not tables:
        return

    if "Summary" in workbook.sheetnames:
        summary_ws = workbook["Summary"]
        workbook.remove(summary_ws)

    summary_ws = workbook.create_sheet("Summary", index=2)

    current_row = 1
    legend_meta: Optional[Tuple[int, pd.DataFrame]] = None

    for title, table_df in tables:
        summary_ws.cell(row=current_row, column=1, value=title).font = Font(bold=True, size=12)
        current_row += 1

        if table_df.empty:
            summary_ws.cell(row=current_row, column=1, value="(no data available)")
            current_row += 2
            continue

        header_row = current_row
        for col_idx, column_name in enumerate(table_df.columns, start=1):
            cell = summary_ws.cell(row=header_row, column=col_idx, value=column_name)
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

        for _, row in table_df.iterrows():
            current_row += 1
            for col_idx, column_name in enumerate(table_df.columns, start=1):
                value = row[column_name]
                summary_ws.cell(row=current_row, column=col_idx, value=value)

        current_row += 2

        if title == "Status Colour Legend":
            legend_meta = (header_row, table_df)

    # Auto width
    for column_cells in summary_ws.columns:
        max_length = 0
        column_letter = get_column_letter(column_cells[0].column)
        for cell in column_cells:
            value = cell.value
            if value is None:
                continue
            text = str(value)
            if "\n" in text:
                text = max(text.splitlines(), key=len)
            max_length = max(max_length, len(text))
        summary_ws.column_dimensions[column_letter].width = max(12, min(max_length + 2, 80))

    summary_ws.freeze_panes = "A2"

    if legend_meta:
        header_row, legend_df = legend_meta
        data_start = header_row + 1
        for idx, status in enumerate(legend_df["Status"], start=data_start):
            fill = STATUS_COLORS.get(status)
            if fill:
                summary_ws.cell(row=idx, column=1).fill = fill

    # Align all cells to top-left with wrap
    for row in summary_ws.iter_rows(min_row=2, max_row=summary_ws.max_row, min_col=1, max_col=summary_ws.max_column):
        for cell in row:
            if cell.alignment is None or not cell.alignment.wrap_text:
                cell.alignment = Alignment(vertical="top", wrap_text=True)


def _populate_custom_sheet(
    workbook,
    sheet_name: str,
    tables: List[Tuple[str, pd.DataFrame]],
) -> None:
    if sheet_name in workbook.sheetnames:
        ws_existing = workbook[sheet_name]
        workbook.remove(ws_existing)
    ws = workbook.create_sheet(sheet_name)
    current_row = 1
    for title, table_df in tables:
        ws.cell(row=current_row, column=1, value=title).font = Font(bold=True, size=12)
        current_row += 1

        if table_df.empty:
            ws.cell(row=current_row, column=1, value="(no data available)")
            current_row += 2
            continue

        header_row = current_row
        for col_idx, column_name in enumerate(table_df.columns, start=1):
            cell = ws.cell(row=header_row, column=col_idx, value=column_name)
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

        for _, row in table_df.iterrows():
            current_row += 1
            for col_idx, column_name in enumerate(table_df.columns, start=1):
                ws.cell(row=current_row, column=col_idx, value=row[column_name])

        current_row += 2

    # Auto width
    for column_cells in ws.columns:
        max_length = 0
        column_letter = get_column_letter(column_cells[0].column)
        for cell in column_cells:
            value = cell.value
            if value is None:
                continue
            text = str(value)
            if "\n" in text:
                text = max(text.splitlines(), key=len)
            max_length = max(max_length, len(text))
        ws.column_dimensions[column_letter].width = max(12, min(max_length + 2, 80))

    ws.freeze_panes = "A2"

    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
        for cell in row:
            if cell.alignment is None or not cell.alignment.wrap_text:
                cell.alignment = Alignment(vertical="top", wrap_text=True)


def _format_ratio(numerator: Optional[int], denominator: Optional[int]) -> str:
    numerator = int(numerator or 0)
    denominator = int(denominator or 0)
    if denominator <= 0:
        return f"{numerator} / {denominator} (0.0%)"
    pct = (numerator / denominator) * 100
    return f"{numerator} / {denominator} ({pct:.1f}%)"


def _format_percentage(value: Optional[float]) -> str:
    if value is None:
        return "N/A"
    return f"{value * 100:.1f}%" if 0 <= value <= 1 else f"{value:.1f}"


def _summarise_vote_pattern(vote_text: str) -> Tuple[str, int, int]:
    parsed = _parse_vote_counts(vote_text)
    if not parsed:
        return ("(no data)", 0, 0)
    sorted_counts = sorted(parsed, key=lambda item: (-item[1], item[0]))
    top_votes = sorted_counts[0][1]
    second_votes = sorted_counts[1][1] if len(sorted_counts) > 1 else 0
    pattern_repr = " | ".join(f"{label}:{count}" for label, count in sorted_counts)
    return pattern_repr, top_votes, second_votes


def _parse_vote_counts(text: str) -> List[Tuple[str, int]]:
    if not text:
        return []
    counts: List[Tuple[str, int]] = []
    for part in str(text).split(","):
        part = part.strip()
        if not part:
            continue
        match = re.search(r"(Type\s*\d+).*?(\d+)$", part, re.IGNORECASE)
        if match:
            label = match.group(1).title().replace(" ", " ")
            count = int(match.group(2))
            counts.append((label, count))
            continue
        # fallback: look for last colon
        if ":" in part:
            label_text, count_text = part.rsplit(":", 1)
            label = label_text.strip()
            try:
                count = int(count_text.strip())
            except ValueError:
                continue
            counts.append((label, count))
    return counts


def _prepare_article_detail_rows(
    article_df: pd.DataFrame,
    summary_row: pd.Series,
    source_column: str,
    ai_agreement_column: str,
    human_vs_ai_column: str,
    type_summary_column: str,
    base_columns: List[str],
) -> pd.DataFrame:
    """Annotate and order detail rows for a single article."""
    article_df = article_df.copy()
    article_df[MISMATCH_PAIR_COLUMN] = ""
    article_df[DETAIL_NOTE_COLUMN] = ""

    human_type = str(summary_row.get(HUMAN_Q15_TYPE_COLUMN, "") or "").strip()
    ai_type = str(summary_row.get(AI_Q15_TYPE_COLUMN, "") or "").strip()
    mismatch_pair = ""
    if human_type and ai_type and human_type != ai_type:
        mismatch_pair = f"{human_type} → {ai_type}"

    ai_agreement = str(summary_row.get(ai_agreement_column, "") or "").strip()
    human_vs_ai = str(summary_row.get(human_vs_ai_column, "") or "").strip()
    type_summary = str(summary_row.get(type_summary_column, "") or "").strip()

    def _note_for_row(row: pd.Series) -> str:
        source_value = str(row.get(source_column, "") or "")
        if source_value == "human":
            return type_summary or "Human annotated row"
        if "majority-vote" in source_value:
            return human_vs_ai or ai_agreement
        return str(row.get("Analysis_Status", "") or "")

    article_df["_detail_order"] = article_df[source_column].apply(_detail_sort_key)

    for idx, row in article_df.iterrows():
        note = _note_for_row(row)
        article_df.at[idx, DETAIL_NOTE_COLUMN] = note

        source_value = str(row.get(source_column, "") or "")
        if mismatch_pair and (source_value == "human" or "majority-vote" in source_value):
            article_df.at[idx, MISMATCH_PAIR_COLUMN] = mismatch_pair

    article_df = article_df.sort_values("_detail_order", kind="stable").drop(columns="_detail_order")
    article_df = article_df.reset_index(drop=True)

    ordered_columns = [col for col in base_columns if col in article_df.columns]
    remaining_columns = [col for col in article_df.columns if col not in ordered_columns]
    article_df = article_df[ordered_columns + remaining_columns]
    return article_df


def _detail_sort_key(source_value: str):
    text = str(source_value or "")
    if text == "human":
        return (0, 0)
    if "majority-vote" in text:
        return (1, 0)
    match = re.search(r"run(\d+)", text)
    if match:
        try:
            return (2, int(match.group(1)))
        except ValueError:
            return (2, 0)
    return (3, 0)


def _build_question_column_map(columns: pd.Index) -> Dict[int, str]:
    mapping: Dict[int, str] = {}
    for col in columns:
        match = QUESTION_COLUMN_PATTERN.search(str(col))
        if match:
            mapping[int(match.group(1))] = col
    return mapping


def _derive_q15_vote_counter(ai_rows: pd.DataFrame, q15_column: Optional[str]) -> Counter:
    if not q15_column or q15_column not in ai_rows.columns:
        return Counter()
    votes: Counter = Counter()
    for value in ai_rows[q15_column].dropna():
        normalized = _normalize_type_answer(str(value))
        if normalized:
            votes[normalized] += 1
    return votes


def _format_vote_counter(counter: Counter) -> str:
    if not counter:
        return ""
    sorted_counts = sorted(counter.items(), key=lambda item: (-item[1], item[0]))
    return ", ".join(f"{label}:{count}" for label, count in sorted_counts)


def _normalize_type_answer(answer: str) -> str:
    answer = answer.strip()
    if not answer:
        return ""
    upper = answer.upper()
    match = re.search(r"TYPE\s*([1-4])", upper)
    if match:
        return f"Type {match.group(1)}"
    match = re.search(r"\b([1-4])\b", answer)
    if match:
        return f"Type {match.group(1)}"
    return answer


def _collect_type_metrics(
    ai_rows: pd.DataFrame,
    question_map: Dict[int, str],
) -> Tuple[Dict[int, float], Dict[int, float]]:
    if ai_rows.empty:
        return {}, {}

    success_mask = ai_rows["Analysis_Status"].astype(str).str.startswith("SUCCESS_")
    success_rows = ai_rows[success_mask]
    if success_rows.empty:
        return {}, {}

    extent_avgs: Dict[int, float] = {}
    confidence_avgs: Dict[int, float] = {}

    for type_id, question_refs in Config.TYPE_QUESTION_GROUPS.items():
        confidence_col = question_map.get(question_refs["confidence"])

        if confidence_col and confidence_col in success_rows.columns:
            confidence_values = []
            for value in success_rows[confidence_col]:
                parsed = ResponseParser.extract_confidence_value(value)
                if parsed is not None:
                    confidence_values.append(parsed)
            if confidence_values:
                confidence_avgs[type_id] = sum(confidence_values) / len(confidence_values)

    return extent_avgs, confidence_avgs


def _determine_article_status(
    human_vs_ai: str,
    human_vs_consensus: str,
    has_success: bool,
    *,
    ai_agreement_label: str,
    majority_available: bool,
    vote_counts_text: str,
    success_count: int,
    total_runs: int,
    ai_majority_value: str,
) -> str:
    normalized_ai = (human_vs_ai or "").strip().lower()
    normalized_consensus = (human_vs_consensus or "").strip().lower()
    agreement_norm = (ai_agreement_label or "").strip().lower()
    vote_counts_norm = (vote_counts_text or "").strip().lower()
    majority_value_norm = (ai_majority_value or "").strip().lower()

    if not has_success or success_count <= 0:
        return "Technical_Failure"

    if "mismatch" in normalized_ai or "mismatch" in normalized_consensus:
        return "Contradiction"

    tie_condition = (
        "split consensus" in agreement_norm
        or "[tie]" in majority_value_norm
        or "[tie]" in vote_counts_norm
        or vote_counts_norm.startswith("tie")
        or " tie" in vote_counts_norm
    )
    if tie_condition:
        return "Ambiguous_Tie"

    poor_coverage_tokens = [
        "no ai majority",
        "ai majority missing",
        "human missing",
        "human unclassified",
        "majority vote disabled",
        "majority vote not applicable",
        "majority vote unavailable",
        "insufficient data",
        "no q15 data",
        "consensus unavailable",
        "consensus missing",
        "consensus unclear",
    ]
    if (
        not majority_available
        or "plurality" in agreement_norm
        or "insufficient data" in agreement_norm
        or "no q15 data" in agreement_norm
        or any(token in normalized_ai for token in poor_coverage_tokens)
        or any(token in normalized_consensus for token in poor_coverage_tokens)
        or total_runs <= 1
    ):
        return "Ambiguous_PoorCoverage"

    ai_match = normalized_ai.startswith("match") and "missing" not in normalized_ai
    if ai_match:
        if "unanimous" in agreement_norm or "strong majority" in agreement_norm:
            return "Pass_Strong"
        if "simple majority" in agreement_norm:
            return "Pass_Weak"
        if "majority" in agreement_norm and "plurality" not in agreement_norm:
            return "Pass_Weak"

    return "Ambiguous_PoorCoverage"


def _format_all_results_sheet(ws, article_id_column: str, article_status_column: str) -> None:
    if ws is None:
        return

    _apply_common_sheet_formatting(ws, freeze_cell="D2", article_status_column=article_status_column)
    _apply_article_separators(ws, article_id_column)


def _format_summary_sheet(ws, article_status_column: str) -> None:
    if ws is None:
        return

    _apply_common_sheet_formatting(ws, freeze_cell="C2", article_status_column=article_status_column)


def _apply_common_sheet_formatting(
    ws,
    *,
    freeze_cell: str,
    article_status_column: str,
) -> None:
    if ws.max_row >= 1 and ws.max_column >= 1:
        ws.auto_filter.ref = f"A1:{get_column_letter(ws.max_column)}{ws.max_row}"

    ws.freeze_panes = freeze_cell

    header_alignment = Alignment(wrap_text=True, horizontal="center", vertical="center")
    for cell in ws[1]:
        cell.alignment = header_alignment

    wrap_alignment = Alignment(wrap_text=True, vertical="top")
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
        for cell in row:
            cell.alignment = wrap_alignment

    for column_index, column_cells in enumerate(ws.columns, start=1):
        max_length = 0
        for cell in column_cells:
            value = cell.value
            if value is None:
                continue
            text = str(value)
            if "\n" in text:
                text = max(text.splitlines(), key=len)
            max_length = max(max_length, len(text))
        adjusted_width = max(10, min(max_length + 2, 80))
        ws.column_dimensions[get_column_letter(column_index)].width = adjusted_width

    status_col_idx = _find_column_index(ws, article_status_column)
    if status_col_idx is None:
        return

    for row_idx in range(2, ws.max_row + 1):
        cell = ws.cell(row=row_idx, column=status_col_idx)
        fill = STATUS_COLORS.get(str(cell.value))
        if fill:
            cell.fill = fill


def _apply_article_separators(ws, article_id_column: str) -> None:
    id_col_idx = _find_column_index(ws, article_id_column)
    if id_col_idx is None:
        return

    thick_side = Side(style="medium", color="000000")
    previous_value = None
    for row_idx in range(2, ws.max_row + 1):
        current_value = ws.cell(row=row_idx, column=id_col_idx).value
        if row_idx == 2 or current_value != previous_value:
            for col_idx in range(1, ws.max_column + 1):
                cell = ws.cell(row=row_idx, column=col_idx)
                existing_border = cell.border or Border()
                cell.border = Border(
                    left=existing_border.left,
                    right=existing_border.right,
                    top=thick_side,
                    bottom=existing_border.bottom,
                )
        previous_value = current_value


def _find_column_index(ws, column_name: str) -> Optional[int]:
    for idx, cell in enumerate(ws[1], start=1):
        if cell.value == column_name:
            return idx
    return None


def _print_status_report(status_map: Mapping[str, str], run_stats: Mapping[str, Counter]) -> None:
    if not status_map:
        return

    print("\n📌 Article status breakdown:")
    status_counter = Counter(status_map.values())
    for status, count in status_counter.most_common():
        print(f"  - {status}: {count}")

    technical_articles = [key for key, value in status_map.items() if value == "Technical_Failure"]
    if not technical_articles:
        return

    print("\n  Technical issues by article:")
    for article_id in technical_articles:
        counter = run_stats.get(article_id, Counter())
        issues = ", ".join(f"{label}:{count}" for label, count in counter.items()) or "No AI runs recorded"
        print(f"    • Article {article_id}: {issues}")
